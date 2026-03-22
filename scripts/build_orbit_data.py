import os
import sys
import re
import csv
import datetime as dt
from typing import List, Tuple, Optional

import pandas as pd
import openpyxl
from dateutil import parser as date_parser


# ============================================================
# Project O.R.B.I.T — Data Builder (Excel -> Clean CSVs)
# - Reads academic_command_centre_v3.xlsx (formatted workbook)
# - Exports Power BI-ready CSVs into .\data
# - Creates templates for missing logs/config in .\logs
# ============================================================


def project_root_from_script() -> str:
    script_path = os.path.abspath(__file__)
    script_dir = os.path.dirname(script_path)
    root = os.path.abspath(os.path.join(script_dir, ".."))
    return root


def ensure_dir(path: str) -> None:
    if not os.path.exists(path):
        os.makedirs(path, exist_ok=True)


def safe_float(x):
    try:
        if x is None:
            return None
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip().replace("%", "")
        return float(s)
    except Exception:
        return None


def parse_table_range(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    col_map: List[Tuple[str, str]],
    start_row: Optional[int] = None,
    max_rows: int = 500,
) -> pd.DataFrame:
    """
    Extracts a table using explicit column letters and desired output field names.
    Stops when ALL mapped columns are blank for a row.
    """
    if start_row is None:
        start_row = header_row + 1

    def col_idx(letter: str) -> int:
        return openpyxl.utils.column_index_from_string(letter)

    rows = []
    for r in range(start_row, start_row + max_rows):
        all_blank = True
        for col_letter, _field in col_map:
            v = ws.cell(r, col_idx(col_letter)).value
            if v not in (None, ""):
                all_blank = False
                break
        if all_blank:
            break

        row_vals = []
        for col_letter, _field in col_map:
            row_vals.append(ws.cell(r, col_idx(col_letter)).value)
        rows.append(row_vals)

    df = pd.DataFrame(rows, columns=[field for _c, field in col_map])
    return df


def write_csv(df: pd.DataFrame, out_path: str) -> None:
    df.to_csv(out_path, index=False, encoding="utf-8-sig")


def extract_research_progress(ws: openpyxl.worksheet.worksheet.Worksheet) -> pd.DataFrame:
    """
    Research Hub sheet stores:
    - Paper 1 completion: D4 (fraction like 0.6)
    - Paper 2 completion: J4 (fraction like 0.0)
    """
    p1 = safe_float(ws["D4"].value)
    p2 = safe_float(ws["J4"].value)

    # Normalize if someone typed 60 instead of 0.60
    if p1 is not None and p1 > 1.0:
        p1 = p1 / 100.0
    if p2 is not None and p2 > 1.0:
        p2 = p2 / 100.0

    df = pd.DataFrame(
        [{
            "paper1_completion_frac": p1 if p1 is not None else 0.0,
            "paper2_completion_frac": p2 if p2 is not None else 0.0,
        }]
    )
    return df


def extract_finance_summary(ws: openpyxl.worksheet.worksheet.Worksheet) -> pd.DataFrame:
    """
    Finance sheet stores summary values:
    A5 = total spent
    D5 = remaining budget
    G5 = weekly allowance
    J5 = savings target
    M5 = saved so far
    """
    spent = safe_float(ws["A5"].value) or 0.0
    remaining = safe_float(ws["D5"].value) or 0.0
    allowance = safe_float(ws["G5"].value) or 0.0
    savings_target = safe_float(ws["J5"].value) or 0.0
    saved = safe_float(ws["M5"].value) or 0.0

    total_budget = spent + remaining if (spent is not None and remaining is not None) else None
    df = pd.DataFrame([{
        "total_budget_yen": total_budget if total_budget is not None else 0.0,
        "total_spent_yen": spent,
        "remaining_budget_yen": remaining,
        "weekly_allowance_yen": allowance,
        "savings_target_yen": savings_target,
        "saved_so_far_yen": saved,
    }])
    return df


def clean_finance_transactions(df: pd.DataFrame) -> pd.DataFrame:
    # Drop rows that are fully empty / placeholders
    df2 = df.copy()
    df2["wk"] = df2["wk"].apply(lambda x: None if str(x).strip() in ("None", "nan", "") else x)
    df2 = df2[~(df2["wk"].isna() & df2["description"].isna() & df2["category"].isna())].copy()

    # Ensure numeric
    for col in ["amount_yen", "running_bal", "pct_of_budget"]:
        if col in df2.columns:
            df2[col] = df2[col].apply(lambda x: safe_float(x) if x not in (None, "") else None)

    # If pct_of_budget is a fraction 0-1 keep; if 0-100 convert
    if "pct_of_budget" in df2.columns:
        df2["pct_of_budget"] = df2["pct_of_budget"].apply(
            lambda v: (v / 100.0) if (v is not None and v > 1.0) else v
        )

    return df2


def create_orbit_config_template(out_path: str) -> None:
    """
    Config drives Trajectory + Drift thresholds in Power BI.
    You can edit values later without touching Power BI measures.
    """
    # Default assumptions (edit later if needed)
    rows = [
        {"key": "student_name", "value": "TANAKA ALEX MBENDANA"},
        {"key": "semester_start_date", "value": "2026-03-02"},
        {"key": "semester_end_date", "value": "2026-06-21"},
        {"key": "attendance_threshold_frac", "value": "0.85"},
        {"key": "weekly_study_baseline_hours", "value": "25"},
        {"key": "budget_overspend_weekly_allowance_frac", "value": "1.00"},
        {"key": "research_stagnant_days", "value": "14"},
    ]
    df = pd.DataFrame(rows)
    df.to_csv(out_path, index=False, encoding="utf-8-sig")


def create_study_log_template(out_path: str) -> None:
    """
    Intelligence domain input (weekly study totals).
    You will update this weekly. Power BI will consume it.
    """
    # Create 8 empty weeks template starting from semester start (can be edited)
    rows = []
    start = dt.date(2026, 3, 2)
    for i in range(1, 9):
        week_start = start + dt.timedelta(days=(i - 1) * 7)
        rows.append({
            "week_number": i,
            "week_start_date": week_start.isoformat(),
            "study_hours_total": "",
            "notes": "",
        })
    df = pd.DataFrame(rows)
    df.to_csv(out_path, index=False, encoding="utf-8-sig")


def compute_kpis(
    df_attendance: pd.DataFrame,
    df_research_progress: pd.DataFrame,
    df_fin_summary: pd.DataFrame,
    df_study_log: Optional[pd.DataFrame],
    config_df: pd.DataFrame,
) -> pd.DataFrame:
    # Attendance KPI
    attended_sum = safe_float(df_attendance["attended"].sum()) if "attended" in df_attendance.columns else 0.0
    total_sum = safe_float(df_attendance["total"].sum()) if "total" in df_attendance.columns else 0.0
    attendance_overall = (attended_sum / total_sum) if total_sum and total_sum > 0 else 0.0

    # Research KPI: (Paper1 + Paper2)/2
    p1 = float(df_research_progress.loc[0, "paper1_completion_frac"]) if not df_research_progress.empty else 0.0
    p2 = float(df_research_progress.loc[0, "paper2_completion_frac"]) if not df_research_progress.empty else 0.0
    research_overall = (p1 + p2) / 2.0

    # Finance KPI
    total_budget = float(df_fin_summary.loc[0, "total_budget_yen"]) if not df_fin_summary.empty else 0.0
    remaining_budget = float(df_fin_summary.loc[0, "remaining_budget_yen"]) if not df_fin_summary.empty else 0.0
    budget_remaining_frac = (remaining_budget / total_budget) if total_budget and total_budget > 0 else 0.0

    # Weekly study total KPI (latest non-empty)
    latest_study_hours = None
    if df_study_log is not None and "study_hours_total" in df_study_log.columns:
        tmp = df_study_log.copy()
        tmp["study_hours_total_num"] = tmp["study_hours_total"].apply(lambda x: safe_float(x))
        tmp2 = tmp.dropna(subset=["study_hours_total_num"])
        if not tmp2.empty:
            latest_study_hours = float(tmp2.iloc[-1]["study_hours_total_num"])
    if latest_study_hours is None:
        latest_study_hours = 0.0

    # Semester completion and days remaining computed from config dates vs today
    def get_cfg(key: str, default: str) -> str:
        hit = config_df[config_df["key"] == key]
        return str(hit.iloc[0]["value"]) if not hit.empty else default

    start_date = dt.date.fromisoformat(get_cfg("semester_start_date", "2026-03-02"))
    end_date = dt.date.fromisoformat(get_cfg("semester_end_date", "2026-06-21"))
    today = dt.date.today()

    duration = (end_date - start_date).days
    elapsed = (today - start_date).days
    if duration <= 0:
        semester_completion = 0.0
    else:
        semester_completion = max(0.0, min(1.0, elapsed / duration))

    days_remaining = max(0, (end_date - today).days)

    # Output single-row KPI table
    kpi = pd.DataFrame([{
        "kpi_date": today.isoformat(),
        "overall_attendance_frac": attendance_overall,
        "research_progress_frac": research_overall,
        "budget_remaining_frac": budget_remaining_frac,
        "weekly_study_hours": latest_study_hours,
        "semester_completion_frac": semester_completion,
        "days_remaining": days_remaining,
    }])
    return kpi


def main():
    root = project_root_from_script()
    os.chdir(root)

    xlsx_path = os.path.join(root, "academic_command_centre_v3.xlsx")
    if not os.path.exists(xlsx_path):
        print(f"ERROR: Cannot find Excel file:\n  {xlsx_path}")
        sys.exit(1)

    data_dir = os.path.join(root, "data")
    logs_dir = os.path.join(root, "logs")
    ensure_dir(data_dir)
    ensure_dir(logs_dir)

    print("Project root:", root)
    print("Reading workbook:", xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # 1) Analytics — Course Attendance table
    ws_analytics = wb["📊 Analytics"]
    df_course_att = parse_table_range(
        ws_analytics,
        header_row=3,
        col_map=[("A", "course"), ("B", "attended"), ("C", "total"), ("D", "attendance_pct_raw"), ("E", "remaining")],
    )

    # Clean % column to fraction if needed
    df_course_att["attendance_pct_frac"] = df_course_att["attendance_pct_raw"].apply(lambda v: safe_float(v))
    df_course_att["attendance_pct_frac"] = df_course_att["attendance_pct_frac"].apply(
        lambda v: (v / 100.0) if (v is not None and v > 1.0) else v
    )

    out_course_att = os.path.join(data_dir, "course_attendance.csv")
    write_csv(df_course_att, out_course_att)
    print("Wrote:", out_course_att)

    # 2) Research Hub — progress + section tables
    ws_research = wb["📄 Research Hub"]
    df_research_progress = extract_research_progress(ws_research)
    out_research_progress = os.path.join(data_dir, "research_progress.csv")
    write_csv(df_research_progress, out_research_progress)
    print("Wrote:", out_research_progress)

    df_paper1 = parse_table_range(
        ws_research,
        header_row=7,
        col_map=[("B", "section"), ("C", "status"), ("D", "target_date"), ("F", "notes")],
    )
    df_paper1["paper"] = "Paper 1"
    out_paper1 = os.path.join(data_dir, "research_paper1_sections.csv")
    write_csv(df_paper1, out_paper1)
    print("Wrote:", out_paper1)

    df_paper2 = parse_table_range(
        ws_research,
        header_row=7,
        col_map=[("H", "section"), ("I", "status"), ("J", "target_date"), ("L", "notes")],
    )
    df_paper2["paper"] = "Paper 2"
    out_paper2 = os.path.join(data_dir, "research_paper2_sections.csv")
    write_csv(df_paper2, out_paper2)
    print("Wrote:", out_paper2)

    # 3) Finance — summary + transactions
    ws_fin = wb["💰 Finance"]
    df_fin_summary = extract_finance_summary(ws_fin)
    out_fin_summary = os.path.join(data_dir, "finance_summary.csv")
    write_csv(df_fin_summary, out_fin_summary)
    print("Wrote:", out_fin_summary)

    df_fin_tx = parse_table_range(
        ws_fin,
        header_row=7,
        col_map=[("A", "wk"), ("B", "description"), ("C", "category"),
                 ("D", "amount_yen"), ("E", "running_bal"), ("F", "pct_of_budget"), ("G", "notes")],
    )
    df_fin_tx = clean_finance_transactions(df_fin_tx)
    out_fin_tx = os.path.join(data_dir, "finance_transactions.csv")
    write_csv(df_fin_tx, out_fin_tx)
    print("Wrote:", out_fin_tx)

    # 4) Config + Study Log templates (if missing)
    config_path = os.path.join(logs_dir, "orbit_config.csv")
    if not os.path.exists(config_path):
        create_orbit_config_template(config_path)
        print("Created template:", config_path)
    else:
        print("Found existing:", config_path)

    study_log_path = os.path.join(logs_dir, "study_log.csv")
    if not os.path.exists(study_log_path):
        create_study_log_template(study_log_path)
        print("Created template:", study_log_path)
    else:
        print("Found existing:", study_log_path)

    config_df = pd.read_csv(config_path)
    try:
        study_df = pd.read_csv(study_log_path)
    except Exception:
        study_df = None

    # 5) KPI single-row table
    df_kpi = compute_kpis(
        df_attendance=df_course_att.rename(columns={"attended": "attended", "total": "total"}),
        df_research_progress=df_research_progress,
        df_fin_summary=df_fin_summary,
        df_study_log=study_df,
        config_df=config_df,
    )
    out_kpi = os.path.join(data_dir, "orbit_kpis.csv")
    write_csv(df_kpi, out_kpi)
    print("Wrote:", out_kpi)

    print("\nDONE. Power BI should connect to:")
    print(" - data\\*.csv")
    print(" - logs\\orbit_config.csv")
    print(" - logs\\study_log.csv")


if __name__ == "__main__":
    main()